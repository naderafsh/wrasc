from locale import format_string
from os import path, rename
import yaml as ym
import time
import regex as re
from xlrd import Book as xlrdbook
from openpyxl import workbook as openpybook


jira_key_regex = r"[A-Z]+\-\d+(?=\s)"
param_regex = r"[a-zA-Z]+_\w*"
req_code_regex = r"[A-Z]+_\w*"


def set_test_params(tst, motor_id):
    """setup the test parameters

    Args:
        tst ([type]): [description]
    """

    tst[motor_id]["mres"] = (
        1
        / tst[motor_id]["fullsteps_per_rev"]
        / tst[motor_id]["micro_steps"]
        * tst[motor_id]["overall_egu_per_rev"]
    )

    if "motor_unit_per_rev" in tst[motor_id]:
        step_res = tst[motor_id]["motor_unit_per_rev"]
    else:
        step_res = tst[motor_id]["mres"]

    enc_res = tst[motor_id]["encoder_res"]
    tst[motor_id]["smalljog_steps"] = tst[motor_id]["smalljog_egu"] / step_res
    tst[motor_id]["bigjog_steps"] = tst[motor_id]["bigjog_egu"] / step_res

    tst[motor_id]["jog_step_ratio"] = (
        tst[motor_id]["bigjog_egu"] / tst[motor_id]["smalljog_egu"]
    )
    tst[motor_id]["HomeOffset"] = tst[motor_id]["HomeOffset_EGU"] / step_res
    tst[motor_id]["attackpos_enc"] = (
        tst[motor_id]["attackpos_egu"] + tst[motor_id]["HomeOffset_EGU"]
    ) / enc_res
    tst[motor_id]["JogSpeed"] = tst[motor_id]["JogSpeed_EGU"] / step_res / 1000
    tst[motor_id]["HomeVel"] = tst[motor_id]["HomeVel_EGU"] / step_res / 1000

    tst[motor_id]["fullrange_steps"] = tst[motor_id]["fullrange_egu"] / step_res

    tst[motor_id]["clearance_enc"] = tst["clearance_egu"] / enc_res

    return tst

    # it is possible to use multiple gpascii channels,
    # but we don't have a reason to do so, yet!


class ShortHand:
    """
    This class auto completes shorthanded text messages/naes/references based on the recent activity.
    Based on the preset format, it tries to guess
    and complete shorthanded leading ( and maybe trailing ) text.
    """

    # expression = r"\([^\(]*<([^<]*)>[^\(]*\)"
    # expression = re.compile(expression)

    # reversed = re.sub(expression, partial(_group_replacer, data), string)
    short_text = ...  # type: str

    def __init__(
        self, group_formats: list, ditto_char="/", pre_dittos=False, post_dittos=False
    ) -> None:

        self.format_list = group_formats
        self.ditto_char = ditto_char
        self.pre_dittos = pre_dittos
        self.post_dittos = post_dittos

        self.full_expression = ""
        for group in self.format_list:
            self.full_expression += "(" + group[0] + ")*" + group[1]

        self.template = re.compile(self.full_expression)
        self.text_groups = [None] * len(group_formats)

    def long(self, short_text: str):
        """generates long form from short hand input
            by filling in the blanks using latest inputs

        Args:
            short_text (str): [description]

        Returns:
            [type]: [complete long form inferred from history]
        """

        self.short_text = short_text
        self._decompose()
        return self._compose()

    def _decompose(self):

        # parse the new text input:
        match = re.search(self.full_expression, self.short_text)
        match_group = match.group()
        match_groups = re.findall(self.full_expression, self.short_text)
        for mg in match_groups:
            if any(mg):
                match_group = list(mg)
                # the first non-empty group is accepted
                break

        pre_text = True
        post_text = False
        chars_found_count = 0
        dittos_found_count = 0
        for i, old_text in enumerate(self.text_groups):

            if match_group[i]:
                pre_text = False
                chars_found_count += len(match_group[i])
                self.text_groups[i] = match_group[i]
            elif old_text:
                post_text = not pre_text
                expecting_dittos = (self.pre_dittos and pre_text) or (
                    self.post_dittos and post_text
                )
                expected_position = dittos_found_count + chars_found_count

                if (len(self.short_text) > expected_position) and (
                    self.short_text[expected_position] == self.ditto_char
                ):
                    # a missing word... is there a ditto here
                    dittos_found_count += 1
                else:
                    # blank in this placeholder
                    if expecting_dittos:
                        # this is an error, because we haven't found no text and no dittos
                        raise RuntimeError(
                            f"Dittos missing at position {expected_position} of {self.short_text}"
                        )

            else:
                # both are blank, we are confused!!
                raise RuntimeError(
                    f"No pretext to complete {self.short_text}, pretext is {self.text_groups} "
                )

    def _compose(self):
        # compose the long form
        long_text = ""  # type: str
        # compose the full length output from self.text_fields
        for i, group in enumerate(self.format_list):
            long_text += self.text_groups[i] + group[1]

        return long_text


def xlrd_sheet_to_dict(wb, sheet_name, heading_row, first_col, key_col):
    """makes a dict for sheet specified by sheet_name
    keys shall be in rows, (one column)

    excel workbook (.xls or .xlsx) into object wb


    Args:
        wb (xlrd.Book): workbook source object
        sheet_name (str): sheet name
        heading_row (int, optional): where heading row is found (Excel row 1 is index 0). Defaults to 2.
        first_col (int, optional): leftmost column of the table (Excel column 'A' is index 0). Defaults to 2.
        key_col (int, optional): dict keys are picked at thios column (Excel column 'A' is index 0). Defaults to None.

    Returns:
        [dict]: [description]
    """

    assert isinstance(wb, xlrdbook)

    sheet = wb.sheet_by_name(sheet_name)
    sheet_dict = dict()
    for row in range(heading_row + 1, sheet.nrows):

        row_dict = dict()
        for col in range(first_col, sheet.ncols):
            value = sheet.cell(row, col).value
            key = str(sheet.cell(heading_row, col).value).strip()
            try:
                value = str(value).strip()
            except ValueError:
                pass

            if col == key_col:
                req_key = value
                continue

            row_dict.update({key: value})

        sheet_dict.update({req_key: row_dict})

    return sheet_dict


def opxl_sheet_to_dict(wb, sheet_name, heading_row=2, first_col=2, key_col=None):

    assert isinstance(wb, openpybook)
    sheet = wb[sheet_name]
    reqs = dict()
    for row in range(heading_row + 1, sheet.max_row):

        req = dict()
        for col in range(first_col, sheet.max_column):
            value = sheet.cell(row + 1, col + 1).value
            key = str(sheet.cell(heading_row + 1, col + 1).value).strip()
            try:
                value = str((value)).strip()
            except ValueError:
                pass

            if col == key_col:
                req_key = value
                continue

            req.update({key: value})

        reqs.update({req_key: req})

    return reqs


def avoid_overwrite(filepath):

    n_copies = 0
    while path.exists(filepath):
        name, ext = path.splitext(filepath)
        modif_time_str = time.strftime(
            "%y%m%d_%H%M", time.localtime(path.getmtime(filepath)),
        )
        n_copies_str = f"({n_copies})" if n_copies > 0 else ""
        try:
            rename(
                filepath, f"{name}_{modif_time_str}{n_copies_str}{ext}",
            )
        except FileExistsError:
            # next copy... never overwrite
            n_copies += 1

    return filepath


def parse_params(text):
    """parse text and return parameters in ccc_ccc format split via ","

    Args:
        text ([str]): [text which includes params]

    Returns:
        [list]: [of strings]
    """

    params_list = text if isinstance(text, list) else re.findall(param_regex, text)

    return params_list if len(params_list) > 0 else []


def query_field(any_table, parameters_field):
    """
    finds parameters listed in field "parameters_field" in table "any_table"
    returns the collection of all paameters in parameter_set
    also replaces the original parameter text with parsed list
    """
    parameter_set = set()

    for _record, _fields in any_table.items():
        _fields: dict
        new_parameters = parse_params(_fields[parameters_field])

        parameter_set = parameter_set.union(set(new_parameters))

        # DONE remove the side effect
        # _fields[parameters_field] = new_parameters

    return parameter_set


def dump_obj(myobj, yaml_dump_file):

    with open(yaml_dump_file, "w+") as f:
        try:
            ym.safe_dump(myobj, f, default_flow_style=False)
            success = True
        except Exception:
            success = False
        finally:
            f.close()
            return success


def undump_obj(obj_str, yaml_dump_path):

    with open(path.join(yaml_dump_path, obj_str + ".yaml"), "r") as f:
        myobjects = ym.safe_load(f)
        f.close()

    return myobjects


def dump_class_objs(yourself, obj_str_list, yaml_dump_path):
    """dumps all dicts in the class to corresponding yaml files in requested folder
    yaml file names follow member names

    Args:
        yourself (class): []
        obj_str_list (str): list of members to be dumped
        yaml_dump_path (str): path of dump folder

    Returns:
        [type]: [description]
    """

    list_of_errors = []
    for obj_str in obj_str_list:

        # yaml doesn't know how to represent objects. so to dump dicts of objects, only dump

        if not dump_obj(
            eval("yourself." + obj_str), path.join(yaml_dump_path, obj_str + ".yaml"),
        ):
            list_of_errors.append(obj_str)
    return list_of_errors


def time_stamp(filename):

    create_time_str = time.strftime("%y%m%d_%H%M", time.localtime())
    name, ext = path.splitext(filename)

    return f"{name}_{create_time_str}{ext}"


if __name__ == "__main__":
    pass
